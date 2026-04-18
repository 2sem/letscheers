#!/usr/bin/env python3
import json
import os
import subprocess

import openpyxl

XLSX_PATH = "Projects/App/Resources/Excel/cheers.xlsx"
SHEET_NAME = "기본_건배사"
PR_BODY_PATH = ".github/scripts/pr_body.md"


def get_last_modified_date() -> str:
    result = subprocess.run(
        ["git", "log", "-1", "--format=%ci", f"{XLSX_PATH}.secret"],
        capture_output=True, text=True,
    )
    date = result.stdout.strip()
    return date[:10] if date else "2024-01-01"


def read_column_map(ws) -> dict:
    return {cell.value: cell.column for cell in ws[1] if cell.value}


def read_existing_toasts(ws, col_map: dict) -> tuple:
    no_col = col_map.get("no")
    title_col = col_map.get("건배사")
    contents_col = col_map.get("의미")
    category_col = col_map.get("분류")

    existing, max_no = [], 0
    for row in ws.iter_rows(min_row=2):
        title_val = row[title_col - 1].value if title_col else None
        if not title_val:
            break
        no_val = row[no_col - 1].value if no_col else None
        no = int(no_val) if no_val else 0
        existing.append({
            "no": no,
            "title": title_val,
            "contents": (row[contents_col - 1].value if contents_col else "") or "",
            "category": (row[category_col - 1].value if category_col else "") or "",
        })
        max_no = max(max_no, no)

    return existing, max_no


def search_new_toasts(categories: list, since_date: str) -> list:
    categories_str = ", ".join(categories)
    prompt = (
        f"Search the web for new Korean 건배사 (drinking toast phrases) popular since {since_date}. "
        f"Categories: {categories_str}. "
        f"Find 10 new ones spread across categories. "
        f"Return JSON array only: "
        f'[{{"title":"건배사","contents":"의미","category":"분류"}}]'
    )

    result = subprocess.run(
        ["claude", "-p", prompt, "--output-format", "json"],
        capture_output=True, text=True,
    )

    try:
        data = json.loads(result.stdout)
        text = data.get("result", "")
    except (json.JSONDecodeError, KeyError):
        text = result.stdout

    start = text.find("[")
    end = text.rfind("]") + 1
    if start != -1 and end > start:
        try:
            return json.loads(text[start:end])
        except json.JSONDecodeError:
            pass

    return []


def append_toast(ws, col_map: dict, no: int, title: str, contents: str, category: str):
    next_row = ws.max_row + 1
    if no_col := col_map.get("no"):
        ws.cell(row=next_row, column=no_col).value = no
    if title_col := col_map.get("건배사"):
        ws.cell(row=next_row, column=title_col).value = title
    if contents_col := col_map.get("의미"):
        ws.cell(row=next_row, column=contents_col).value = contents
    if category_col := col_map.get("분류"):
        ws.cell(row=next_row, column=category_col).value = category


def write_pr_body(new_toasts: list, since_date: str):
    lines = [
        "## 건배사 데이터 업데이트",
        "",
        f"**{since_date}** 이후 새로운 건배사 **{len(new_toasts)}개** 추가",
        "",
        "| 건배사 | 의미 | 분류 |",
        "|--------|------|------|",
        *[f"| {t['title']} | {t.get('contents', '')} | {t.get('category', '')} |" for t in new_toasts],
    ]
    with open(PR_BODY_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def set_output(key: str, value: str):
    github_output = os.environ.get("GITHUB_OUTPUT")
    if github_output:
        with open(github_output, "a") as f:
            f.write(f"{key}={value}\n")
    else:
        print(f"OUTPUT: {key}={value}")


def main():
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb[SHEET_NAME]

    col_map = read_column_map(ws)
    existing, max_no = read_existing_toasts(ws, col_map)
    existing_titles = {t["title"] for t in existing}
    categories = sorted({t["category"] for t in existing if t["category"]})

    since_date = get_last_modified_date()
    print(f"Existing: {len(existing)} toasts | Since: {since_date} | Categories: {categories}")

    new_toasts = search_new_toasts(categories, since_date)
    new_toasts = [t for t in new_toasts if t.get("title") and t["title"] not in existing_titles]

    if not new_toasts:
        print("No new toasts found")
        set_output("has_new_toasts", "false")
        return

    print(f"Adding {len(new_toasts)} new toasts")
    for toast in new_toasts:
        max_no += 1
        append_toast(ws, col_map, max_no, toast["title"], toast.get("contents", ""), toast.get("category", ""))

    wb.save(XLSX_PATH)
    write_pr_body(new_toasts, since_date)
    set_output("has_new_toasts", "true")


if __name__ == "__main__":
    main()
